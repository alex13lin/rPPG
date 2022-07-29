# -*- coding: utf-8 -*-
"""
Created on Mon Mar 21 20:02:23 2022

@author: a1016
"""
import datetime
import os
from openpyxl import Workbook,load_workbook
import time
import numpy as np
import zipfile


def zip_dir(path,out_path):
    zf = zipfile.ZipFile('{}.zip'.format(out_path), 'w', zipfile.ZIP_DEFLATED)
   
    for root, dirs, files in os.walk(path):
        for file_name in files:
            zf.write(os.path.join(root, file_name))
            
            
def saving(filename,result_data,selfsubject_num,selflf_data,selfhf_data,selflfhf_data,
           selflf_array,selfhf_array,selflfhf_array,
           selfdlfhf_data,selfSampEn_data,selfApEn_data,
           selfSBP_DBP_data,selfSBP_DBP2_data):
    
    loc_dt = datetime.datetime.today() 
    
    if os.path.isfile(r'D:/fatigue//' + filename + '.xlsx'):
        load_workbook(r'D:/fatigue//' + filename + '.xlsx')
    else:
        Workbook().save(r'D:/fatigue//' + filename + '.xlsx')
        #print("建立新檔案!!!")
        wb = load_workbook(r'D:/fatigue//' + filename + '.xlsx')
        ws = wb.active
        ws.title="基本測試結果"
        result_title=["學號/編號","性別","年齡",
                      "身高","體重","BMI",
                      "坐高","光度",
                      "(實測)心律","(實測)收縮壓","(實測)舒張壓",
                      "(rPPG)心律","(rPPG)收縮壓","(rPPG)舒張壓","眨眼次數",
                      "LF(nu)平均","HF(nu)平均","LF/HF平均",
                      "LF面積(測試)","HF面積(測試)",
                      "第一筆數據時間","數據蒐集總時間","測試時刻"]
        ws.append(result_title)
        wb.create_sheet("LF(nu)")
        wb.create_sheet("HF(nu)")
        wb.create_sheet("LFHF比值")
        wb.create_sheet("LFHF比值變化量")                        
        wb.create_sheet("SampEn")
        wb.create_sheet("ApEn")
        wb.create_sheet("LF面積(測試)")
        wb.create_sheet("HF面積(測試)")
        wb.create_sheet("(rPPG)收縮壓")
        wb.create_sheet("(rPPG)舒張壓")
        wb.save(r'D:/fatigue//' + filename + '.xlsx')
    
    wb = load_workbook(r'D:/fatigue//' + filename + '.xlsx')
    ws1 = wb["基本測試結果"]
    ws2 = wb["LF(nu)"]
    ws3 = wb["HF(nu)"]
    ws4 = wb["LFHF比值"]
    ws5 = wb["LFHF比值變化量"]
    ws6 = wb["SampEn"]
    ws7 = wb["ApEn"]
    ws8 = wb["LF面積(測試)"]
    ws9 = wb["HF面積(測試)"]
    ws10 = wb["(rPPG)收縮壓"]
    ws11 = wb["(rPPG)舒張壓"]
    #寫入基本測試結果
    
    
    
    ws1.append(result_data)
    #print("寫入基本測試結果!!!")
    
    subject_num = [selfsubject_num]
    lf_data = subject_num + selflf_array
    hf_data = subject_num + selfhf_array
    lfhf_data = subject_num + selflfhf_array
    dlfhf_data = subject_num + selfdlfhf_data[1:]
    SampEn_data = subject_num + selfSampEn_data
    ApEn_data = subject_num + selfApEn_data
    
    SBP_DBP_data = subject_num + selfSBP_DBP_data
    SBP_DBP2_data = subject_num + selfSBP_DBP2_data

    #寫入LF(nu)
    
    ws2.append(lf_data)
    
    #寫入HF(nu)
    
    ws3.append(hf_data)
    
    #寫入LF/HF
    
    ws4.append(lfhf_data)
    
    #寫入d(LF/HF)/dt 多數據
    
    ws5.append(dlfhf_data)
    
    #寫入樣本熵
    
    ws6.append(SampEn_data)
    
    #寫入近似熵
    
    ws7.append(ApEn_data)                 
    #print("寫入多數據檔案!!!")             
    
    
    
    
    
    ws10.append(SBP_DBP_data)
    ws11.append(SBP_DBP2_data)
    
    #儲存檔案
    wb.save(r'D:/fatigue//' + filename + '.xlsx')

    
    #壓縮檔案 
                    
    #zip_loc_dt_format = loc_dt.strftime("%Y_%m%d_%H%M")
    #zip_date = str(zip_loc_dt_format)
    #zip_path = r'D:\fatigue'
    #zip_out_path = r'D:\fatigue_records\\'+zip_date
    #zip_dir(zip_path,zip_out_path)
        
    
    
    #上傳檔案
    #if self.subject_upload == True:                        
        #up.upload(filename)
    
    #寫入完成
    
    return True
